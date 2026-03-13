"""Excel xlsx wrapper - 格式转换、批量处理、数据分析增强功能.

提供 mcp-excel-server 之外的额外功能：
- 格式转换：xlsx <-> CSV, xlsx <-> JSON
- 批量处理：批量数据导入导出
- 数据分析：统计分析、图表生成
- 智能分析：数据洞察、趋势分析

Usage:
    python wrapper.py convert <input> <output> [--sheet SHEET]
    python wrapper.py batch-import <directory> <output> [--pattern PATTERN]
    python wrapper.py batch-export <input> <directory> [--format FORMAT]
    python wrapper.py analyze <input> [--sheet SHEET] [--output OUTPUT]
    python wrapper.py insight <input> [--sheet SHEET] [--top N]
"""

from __future__ import annotations

import json
import logging
import sys
from pathlib import Path
from typing import Any

import click
import pandas as pd

logger = logging.getLogger(__name__)


def _read_xlsx(file_path: str, sheet_name: str | None = None) -> pd.DataFrame:
    """读取 xlsx 文件为 DataFrame."""
    kwargs: dict[str, Any] = {"engine": "openpyxl"}
    if sheet_name:
        kwargs["sheet_name"] = sheet_name
    return pd.read_excel(file_path, **kwargs)


def _get_sheet_names(file_path: str) -> list[str]:
    """获取 xlsx 文件的所有工作表名."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


# ── 格式转换 ──────────────────────────────────────────────


def xlsx_to_csv(
    input_path: str,
    output_path: str,
    sheet_name: str | None = None,
) -> str:
    """将 xlsx 转换为 CSV 文件."""
    df = _read_xlsx(input_path, sheet_name)
    df.to_csv(output_path, index=False, encoding="utf-8-sig")
    return output_path


def csv_to_xlsx(
    input_path: str,
    output_path: str,
    sheet_name: str = "Sheet1",
) -> str:
    """将 CSV 转换为 xlsx 文件."""
    df = pd.read_csv(input_path, encoding="utf-8-sig")
    df.to_excel(output_path, sheet_name=sheet_name, index=False, engine="openpyxl")
    return output_path


def xlsx_to_json(
    input_path: str,
    output_path: str,
    sheet_name: str | None = None,
    orient: str = "records",
) -> str:
    """将 xlsx 转换为 JSON 文件."""
    df = _read_xlsx(input_path, sheet_name)
    df.to_json(output_path, orient=orient, force_ascii=False, indent=2)
    return output_path


def json_to_xlsx(
    input_path: str,
    output_path: str,
    sheet_name: str = "Sheet1",
) -> str:
    """将 JSON 转换为 xlsx 文件."""
    df = pd.read_json(input_path)
    df.to_excel(output_path, sheet_name=sheet_name, index=False, engine="openpyxl")
    return output_path


# ── 批量处理 ──────────────────────────────────────────────


def batch_import(
    directory: str,
    output_path: str,
    pattern: str = "*.csv",
) -> dict[str, Any]:
    """批量导入目录中的文件到一个 xlsx 工作簿."""
    dir_path = Path(directory)
    files = sorted(dir_path.glob(pattern))
    if not files:
        raise FileNotFoundError(f"在 {directory} 中未找到匹配 {pattern} 的文件")

    results: dict[str, int] = {}
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for f in files:
            sheet = f.stem[:31]  # Excel 工作表名最长 31 字符
            if pattern.endswith(".csv"):
                df = pd.read_csv(f, encoding="utf-8-sig")
            elif pattern.endswith(".json"):
                df = pd.read_json(f)
            else:
                df = pd.read_excel(f, engine="openpyxl")
            df.to_excel(writer, sheet_name=sheet, index=False)
            results[sheet] = len(df)

    return {"output": output_path, "sheets": results, "total_files": len(files)}


def batch_export(
    input_path: str,
    directory: str,
    fmt: str = "csv",
) -> dict[str, Any]:
    """将 xlsx 中的每个工作表导出为独立文件."""
    dir_path = Path(directory)
    dir_path.mkdir(parents=True, exist_ok=True)

    sheets = _get_sheet_names(input_path)
    results: dict[str, str] = {}

    for sheet in sheets:
        df = pd.read_excel(input_path, sheet_name=sheet, engine="openpyxl")
        if fmt == "csv":
            out = dir_path / f"{sheet}.csv"
            df.to_csv(out, index=False, encoding="utf-8-sig")
        elif fmt == "json":
            out = dir_path / f"{sheet}.json"
            df.to_json(out, orient="records", force_ascii=False, indent=2)
        else:
            raise ValueError(f"不支持的格式: {fmt}，支持 csv/json")
        results[sheet] = str(out)

    return {"input": input_path, "exported": results, "total_sheets": len(sheets)}


# ── 数据分析 ──────────────────────────────────────────────


def analyze_data(
    input_path: str,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """对 xlsx 数据进行统计分析."""
    df = _read_xlsx(input_path, sheet_name)

    numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
    text_cols = df.select_dtypes(include=["object"]).columns.tolist()

    analysis: dict[str, Any] = {
        "shape": {"rows": len(df), "columns": len(df.columns)},
        "columns": list(df.columns),
        "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
        "missing": df.isnull().sum().to_dict(),
        "missing_pct": (df.isnull().sum() / len(df) * 100).round(2).to_dict(),
    }

    if numeric_cols:
        stats = df[numeric_cols].describe().round(4).to_dict()
        analysis["numeric_stats"] = stats

        correlations = df[numeric_cols].corr().round(4)
        analysis["correlations"] = correlations.to_dict()

    if text_cols:
        text_stats: dict[str, Any] = {}
        for col in text_cols:
            vc = df[col].value_counts()
            text_stats[col] = {
                "unique": int(df[col].nunique()),
                "top_values": vc.head(5).to_dict(),
            }
        analysis["text_stats"] = text_stats

    return analysis


def generate_chart(
    input_path: str,
    output_path: str,
    chart_type: str = "bar",
    x_column: str | None = None,
    y_column: str | None = None,
    sheet_name: str | None = None,
    title: str | None = None,
) -> str:
    """生成数据图表并保存为图片."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    df = _read_xlsx(input_path, sheet_name)

    fig, ax = plt.subplots(figsize=(10, 6))

    if chart_type == "bar":
        if x_column and y_column:
            df.plot.bar(x=x_column, y=y_column, ax=ax)
        else:
            df.select_dtypes(include=["number"]).plot.bar(ax=ax)
    elif chart_type == "line":
        if x_column and y_column:
            df.plot.line(x=x_column, y=y_column, ax=ax)
        else:
            df.select_dtypes(include=["number"]).plot.line(ax=ax)
    elif chart_type == "scatter":
        if not x_column or not y_column:
            raise ValueError("scatter 图表需要指定 x_column 和 y_column")
        df.plot.scatter(x=x_column, y=y_column, ax=ax)
    elif chart_type == "histogram":
        if y_column:
            df[y_column].plot.hist(ax=ax, bins=20)
        else:
            df.select_dtypes(include=["number"]).plot.hist(ax=ax, bins=20)
    elif chart_type == "pie":
        if not x_column or not y_column:
            raise ValueError("pie 图表需要指定 x_column（标签）和 y_column（值）")
        df.set_index(x_column)[y_column].plot.pie(ax=ax, autopct="%1.1f%%")
    elif chart_type == "box":
        df.select_dtypes(include=["number"]).plot.box(ax=ax)
    else:
        raise ValueError(f"不支持的图表类型: {chart_type}，支持 bar/line/scatter/histogram/pie/box")

    chart_title = title or f"{Path(input_path).stem} - {chart_type}"
    ax.set_title(chart_title)
    plt.tight_layout()
    fig.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close(fig)

    return output_path


# ── 智能分析 ──────────────────────────────────────────────


def data_insight(
    input_path: str,
    sheet_name: str | None = None,
    top_n: int = 5,
) -> dict[str, Any]:
    """生成数据洞察和趋势分析."""
    df = _read_xlsx(input_path, sheet_name)

    insights: dict[str, Any] = {
        "overview": {
            "total_rows": len(df),
            "total_columns": len(df.columns),
            "memory_usage_mb": round(df.memory_usage(deep=True).sum() / 1024 / 1024, 2),
        },
    }

    # 数据质量评分
    completeness = (1 - df.isnull().sum().sum() / (len(df) * len(df.columns))) * 100
    insights["data_quality"] = {
        "completeness_pct": round(completeness, 2),
        "duplicate_rows": int(df.duplicated().sum()),
        "columns_with_nulls": df.columns[df.isnull().any()].tolist(),
    }

    numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()

    if numeric_cols:
        # 异常值检测（IQR 方法）
        outliers: dict[str, int] = {}
        for col in numeric_cols:
            q1 = df[col].quantile(0.25)
            q3 = df[col].quantile(0.75)
            iqr = q3 - q1
            outlier_mask = (df[col] < q1 - 1.5 * iqr) | (df[col] > q3 + 1.5 * iqr)
            count = int(outlier_mask.sum())
            if count > 0:
                outliers[col] = count
        insights["outliers"] = outliers

        # 高相关性对
        if len(numeric_cols) >= 2:
            corr = df[numeric_cols].corr()
            high_corr: list[dict[str, Any]] = []
            for i in range(len(numeric_cols)):
                for j in range(i + 1, len(numeric_cols)):
                    val = corr.iloc[i, j]
                    if abs(val) > 0.7:
                        high_corr.append({
                            "col1": numeric_cols[i],
                            "col2": numeric_cols[j],
                            "correlation": round(float(val), 4),
                        })
            high_corr.sort(key=lambda x: abs(x["correlation"]), reverse=True)
            insights["high_correlations"] = high_corr[:top_n]

        # 趋势分析（数值列的基本趋势）
        trends: dict[str, str] = {}
        for col in numeric_cols:
            series = df[col].dropna()
            if len(series) >= 3:
                first_half = series.iloc[: len(series) // 2].mean()
                second_half = series.iloc[len(series) // 2 :].mean()
                if second_half > first_half * 1.05:
                    trends[col] = "rising"
                elif second_half < first_half * 0.95:
                    trends[col] = "declining"
                else:
                    trends[col] = "stable"
        insights["trends"] = trends

    # 文本列分析
    text_cols = df.select_dtypes(include=["object"]).columns.tolist()
    if text_cols:
        text_insights: dict[str, Any] = {}
        for col in text_cols[:top_n]:
            vc = df[col].value_counts()
            text_insights[col] = {
                "cardinality": int(df[col].nunique()),
                "top_value": str(vc.index[0]) if len(vc) > 0 else None,
                "top_frequency": int(vc.iloc[0]) if len(vc) > 0 else 0,
                "is_categorical": df[col].nunique() < len(df) * 0.05,
            }
        insights["text_analysis"] = text_insights

    return insights


# ── CLI ───────────────────────────────────────────────────


@click.group()
def cli() -> None:
    """Excel xlsx 包装层 - 格式转换、批量处理、数据分析."""
    logging.basicConfig(level=logging.INFO, format="%(message)s")


@cli.command()
@click.argument("input_path")
@click.argument("output_path")
@click.option("--sheet", default=None, help="指定工作表名")
def convert(input_path: str, output_path: str, sheet: str | None) -> None:
    """格式转换：根据文件扩展名自动判断转换方向."""
    inp = Path(input_path).suffix.lower()
    out = Path(output_path).suffix.lower()

    converters = {
        (".xlsx", ".csv"): lambda: xlsx_to_csv(input_path, output_path, sheet),
        (".csv", ".xlsx"): lambda: csv_to_xlsx(input_path, output_path, sheet or "Sheet1"),
        (".xlsx", ".json"): lambda: xlsx_to_json(input_path, output_path, sheet),
        (".json", ".xlsx"): lambda: json_to_xlsx(input_path, output_path, sheet or "Sheet1"),
    }

    converter = converters.get((inp, out))
    if not converter:
        click.echo(f"不支持的转换: {inp} -> {out}", err=True)
        sys.exit(1)

    result = converter()
    click.echo(json.dumps({"status": "ok", "output": result}, ensure_ascii=False))


@cli.command("batch-import")
@click.argument("directory")
@click.argument("output_path")
@click.option("--pattern", default="*.csv", help="文件匹配模式")
def batch_import_cmd(directory: str, output_path: str, pattern: str) -> None:
    """批量导入目录中的文件到一个 xlsx."""
    result = batch_import(directory, output_path, pattern)
    click.echo(json.dumps(result, ensure_ascii=False, indent=2))


@cli.command("batch-export")
@click.argument("input_path")
@click.argument("directory")
@click.option("--format", "fmt", default="csv", type=click.Choice(["csv", "json"]))
def batch_export_cmd(input_path: str, directory: str, fmt: str) -> None:
    """将 xlsx 每个工作表导出为独立文件."""
    result = batch_export(input_path, directory, fmt)
    click.echo(json.dumps(result, ensure_ascii=False, indent=2))


@cli.command()
@click.argument("input_path")
@click.option("--sheet", default=None, help="指定工作表名")
@click.option("--output", default=None, help="输出图表路径")
@click.option("--chart-type", default=None, help="图表类型: bar/line/scatter/histogram/pie/box")
@click.option("--x", "x_col", default=None, help="X 轴列名")
@click.option("--y", "y_col", default=None, help="Y 轴列名")
def analyze(
    input_path: str,
    sheet: str | None,
    output: str | None,
    chart_type: str | None,
    x_col: str | None,
    y_col: str | None,
) -> None:
    """统计分析，可选生成图表."""
    result = analyze_data(input_path, sheet)
    click.echo(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    if output and chart_type:
        chart_path = generate_chart(
            input_path, output, chart_type, x_col, y_col, sheet,
        )
        click.echo(f"\n图表已保存: {chart_path}")


@cli.command()
@click.argument("input_path")
@click.option("--sheet", default=None, help="指定工作表名")
@click.option("--top", default=5, help="Top N 结果数")
def insight(input_path: str, sheet: str | None, top: int) -> None:
    """智能数据洞察和趋势分析."""
    result = data_insight(input_path, sheet, top)
    click.echo(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    cli()
